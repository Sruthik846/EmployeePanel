from asyncio.windows_events import NULL
from cmath import nan
from django.core.paginator import Paginator,EmptyPage,PageNotAnInteger
from django.http import request,HttpResponse,Http404
from django.shortcuts import redirect, render
from django.views.decorators.csrf import csrf_exempt
from accounts.decorators import allowed_users
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from calendar import monthrange
from django.db.models import Q
from .models import *
from .forms import *
import os
import boto3
import pandas as pd
from openpyxl import load_workbook
import csv ,string
import numpy as np
import textwrap, datetime,time
from accounts.models import EmployeeDetails
from message.models import IFMessages
from openpyxl.styles import PatternFill
import regex as re
from datetime import date
path = os.getcwd()

# ----- Add Month ------- #
@csrf_exempt
@login_required(login_url='loginpage')
@allowed_users(allowed_roles=['admin'])
def addMonth(request):
    if request.method == 'POST':
        month = request.POST.get('month')
        yr = request.POST.get('year')
        if MonthData.objects.filter(year=yr,month=month).exists():
            messages.error(request,'data exists')
        else:
            MonthData.objects.create(year = yr,
                                    month= month)

    monthdatas = MonthData.objects.all().order_by('id').reverse()
    for mt in monthdatas:
        if '-' in mt.month:
            mt.month_name = mt.month.split("-")[-1]
        else:
            mt.month_name = mt.month
    todays_date = date.today()
    context = {'monthdata':monthdatas,'page_title':'MonthWise Details','year':todays_date.year}
    return render(request,'viewmonths.html',context)


@login_required(login_url='loginpage')
@allowed_users(allowed_roles=['admin'])
def delMonth(request,pk):
    months = MonthData.objects.get(id=pk)
    mnth = months.month
    yr = months.year
    months.delete()
    PunchingFile.objects.filter(year=yr,Month=mnth).delete()
    PunchingDatas.objects.filter(month=mnth,year = yr).delete()
    messages.success(request,'data deleted succesfully!')
    return redirect('addmonth')


@login_required(login_url='loginpage')
@allowed_users(allowed_roles=['admin'])
def viewPnFiles(request,year,month):
        
    if '-' in month:
        month_name = month.split("-")[-1]
    else:
        month_name = month
    # print(month,yrSelected)
  
    xlfiles = PunchingFile.objects.filter(Month=month, year=year)
    filedata = []
    for xl in xlfiles:
        # print(xl.file)
        fname = str(xl.file).split('/')[-1]
        xl.file = fname
        filedata.append(xl)
    
    navitems = {'monthwise':'addmonth',}
    context = {'filedata':filedata,'month':month,'year':year, 'month_name':month_name,
                'page_title':'Details'+str(year)+month,'navitems':navitems }
    return render(request,'viewfiles.html', context)


mnth = ''
yr = ''
punchingData = []

@login_required(login_url='loginpage')
@allowed_users(allowed_roles=['admin'])
def PunchTable(request,year,month):
    global mnth, yr,punchingData
    punchingData = PunchingDatas.objects.filter(month=month, year=year).order_by('ifid')
    month_name = month
    if '-' in month:
        month_name = month.split("-")[-1]
    query = request.GET.get('q')
    if query:
        # print(query)
        punchingData = PunchingDatas.objects.filter(month=month, year=year).filter(
            Q(ifid__icontains=query) | Q(name__icontains=query)
        ).distinct()
    paginator = Paginator(punchingData,2000)
    page = request.GET.get('page')
    try:
        rows = paginator.page(page)
    except PageNotAnInteger:
        rows = paginator.page(1)
    except EmptyPage:
        rows = paginator.page(paginator.num_pages)

    navitems = {'monthwise':'addmonth',}
    context = {'punchdata':rows,'month':month,'month_name':month_name,'year':year, 'page_title':month_name+" | "+str(year),'navitems':navitems}
    return render(request,'viewtable.html',context)


att_code = {'FF00FF00':'informed','FF93C47D':'correct','FF00FFFF':'halfday',
        'FFFFFF00':'mistake' , 'FFFCE5CD':'ot','FFFFFFFF':'leave','FFB6D7A8':'correct','FF999999':'leave','00000000':'leave',
        
        }
@login_required(login_url='loginpage')
@allowed_users(allowed_roles=['admin'])
@csrf_exempt
def uploadPunching2(request,year,month):
    if request.method == 'POST':
        punchfile = PunchingFile.objects.create(file = request.FILES.get('punchfile'),
                                    year = year,
                                    Month= month)
        
        xlfile = request.FILES.get('punchfile')
        
        ######## MERGING #################
        f1 = pd.read_excel(xlfile,sheet_name='Total attendance',skiprows=4)
        f2 = pd.read_excel(xlfile,sheet_name='Punching Details',skiprows=8)
        # print(f1.columns)

        f1col = f1.columns
        f2col = f2.columns
        col_included= []
        out_cols = []
        dates=[]
        for i in f1col:
            out_cols.append(i)
        for i in f2col:
            if type(i)== str:
                pass
            else:
                dates.append(i)
                d1 = str(i)
                d2 = d1.split(' ')[0]
                d3 = d2.split('-')[2]
                out_cols.append(d3)
                col_included.append(i)
        col_included.append('IFID')
        # print(dates)
        rowy = 9
        rowx = 5
        alphabest_list = [f"{i}{j}" for i in ["", "A"] for j in string.ascii_uppercase]
        wb1 = load_workbook(xlfile,data_only=True)
        sheet = wb1['Punching Details']
        ws= wb1['Total attendance']
        for i,j in f1.iterrows():
            rowx += 1
            if_id = j['IFID']
            # print(j['Total Attendance of the month'])
            if not np.isnan( j['IFID']) :
                print("Merging...",if_id)
                cell_index=''
                cell_color=''
                #### try except for color errors
                try:
                    for i,rows in f2.iterrows():
                        IFID = rows['IFID']
                        if if_id == IFID:
                            if 'Night Shift' in out_cols:
                                rowy += 1
                                col_value =9
                                colx = 9
                                col =10
                            else:
                                rowy += 1
                                col_value =9
                                colx = 8
                                col =9
                            for days in dates:
                                d1 = str(days)
                                d2 = d1.split(' ')[0]
                                d3 = d2.split('-')[-1]
                                colx += 1
                                col_value +=1
                                col += 1
                                cell_idx = str(alphabest_list[colx]) + str(rowx)
                                ws.cell(row=5,column=col).value = d3
                                cell_index = str(alphabest_list[col_value]) + str(rowy)
                                cell_color = sheet[cell_index].fill.start_color.index
                                # print(cell_index,cell_color,rows[days])
                                ws[cell_idx] = rows[days]
                                # print(cell_idx,rows[days])
                                if len(str(cell_color)) == 1:
                                    cell_color =str(cell_color).zfill(8)
                                fill_cell = PatternFill(patternType='solid', fgColor=cell_color) 
                                ws[cell_idx].fill = fill_cell  
                except:
                    print("error : ",cell_index,cell_color) 
                    messages.error(request,f"Error_______ \n{e} : {cell_idx} {cell_color}  IFID:{if_id}  Plz validate file and upload again!!")
                    return redirect(f'/attendance/viewfile/{year}/{month}')      
        wb1.save(xlfile)
        
        xls = pd.read_excel(xlfile , skiprows=4, sheet_name='Total attendance')
        # print(xls)
        dates_given = []
        for d in xls.columns:
            # print(d)
            if str(d).isnumeric():
                dates_given.append(int(d))
        totaldays = []
        # print(dates_given)
        for item in range(1,32,1):
            totaldays.append(item)
        not_given = list(set(totaldays) - set(dates_given))

        wb = load_workbook(xlfile, data_only = True)
        sh = wb['Total attendance']
        # print(wb.sheetnames)
        sh = wb['Total attendance']
        # print(sh.columns)
        alphaList = [f"{i}{j}" for i in ["", "A"] for j in string.ascii_uppercase]
        rowcount = 5
        for i,row in xls.iterrows():
            # print(row)
            rowcount+=1
            try:
                # check IFID is given;
                if not np.isnan( row['IFID']) :     
                    datedata = {}
                    attdata = {}
                    tozeros  = ['WFO Attendance','OT Attendance','WFH Attendance','Number of punching mistakes','Night Shift','Deductions','Total Attendance of the month']
                    colcount = 2
                    # print(row['Total Attendance of the month'])
                    for i in tozeros:
                        if i in row:
                            colcount+=1
                            if np.isnan(row[i]):
                                attdata[i] = 0
                            else:
                                attdata[i] = row[i]
                                # print(i,row[i])
                        else:
                            attdata[i] = 0
                    for x in not_given: 
                        datedata[x] = 'notgiven__Not available'
                    # print(colcount)
                    for x in dates_given:
                        # print(x)
                        colcount+=1
                        cellindex = str(alphaList[colcount])+str(rowcount)
                        color_in_hex = sh[cellindex].fill.start_color.index
                        att_mode = att_code[color_in_hex]
                        # print(row)
                        y= str(x).zfill(2)
                        if str(row[y])!= "nan":
                            if ( len(str(row[y])) ==8 ) or ( len(str(row[y])) ==5 ):
                                if re.search('[a-zA-Z]', str(row[y])):
                                    # print('__________',att_mode,str(row[y]))
                                    pntime = att_mode+"__"+str(row[y])
                                    datedata[x] = pntime
                                    
                                else:
                                    att_mode = att_code[color_in_hex]
                                    # print('----------',att_mode,str(row[y]))
                                    pntime = str(row[y])[:5]
                                    pntime = att_mode+"__"+pntime
                                    datedata[x] = pntime
                                    
                            else:
                                # print('#########',att_mode,str(row[y]))
                                pntime = att_mode+"__"+str(row[y])
                                datedata[x] = pntime
                                # print(datedata[x])
                                
                        else:
                            # print(str(row[y]))
                            # print('**************',att_mode,str(row[y]))
                            att_mode = att_code[color_in_hex]
                            datedata[x] = att_mode+'__Leave'
                            
                    # update if data exists;
                    existdata = PunchingDatas.objects.filter(ifid = row['IFID'] ,month = month,year = year,)
                    if existdata.exists():
                        print(row['IFID'] ,row['NAME'],'..... exist')   
                    else:

                        print(row['IFID'] ,row['NAME'])
                        PunchingDatas.objects.create( file=punchfile ,
                        ifid = row['IFID'] , name=row['NAME'], month = month,
                        year = year, WFO_attendance = round( attdata['WFO Attendance'],2),
                        WFH_attendance = round( attdata['WFH Attendance'],2), OT_attendance = round( attdata['OT Attendance'],2),
                        Punching_Mistake = round( attdata['Number of punching mistakes']),
                        Night_shift = round( attdata['Night Shift'],2), total_Attendance = round( attdata['Total Attendance of the month'],2) ,
                        Deduction = round( attdata['Deductions'], 2),
                        date_1=datedata[1], date_2=datedata[2], date_3=datedata[3], date_4=datedata[4], date_5=datedata[5], date_6=datedata[6], date_7=datedata[7],
                        date_8=datedata[8], date_9=datedata[9], date_10=datedata[10], date_11=datedata[11], date_12=datedata[12], date_13=datedata[13], date_14=datedata[14],
                        date_15=datedata[15], date_16=datedata[16], date_17=datedata[17], date_18=datedata[18], date_19=datedata[19], date_20=datedata[20], date_21=datedata[21],
                        date_22=datedata[22], date_23=datedata[23], date_24=datedata[24], date_25=datedata[25], date_26=datedata[26], date_27=datedata[27], date_28=datedata[28],
                        date_29=datedata[29], date_30=datedata[30], date_31=datedata[31],
                    )

            except Exception as e:
                print(row['IFID'] ,row['NAME'],cellindex)
                print("error :"+str(e))
                
                PunchingDatas.objects.filter(year = year,month = month,file=punchfile).delete()
                messages.error(request,f"Error_______ \n{e} : {cellindex}  IFID:{int(row['IFID'])}  Plz validate file and upload again!!")
                return redirect(f'viewfile/{year}/{month}')

        return redirect(f'/attendance/viewfile/{year}/{month}')
                
    return redirect(f'/attendance/viewfile/{year}/{month}')



@login_required(login_url='loginpage')
@allowed_users(allowed_roles=['admin'])
def delPunchFile(request,year,month,file_id):
    PunchingFile.objects.filter(id=file_id).delete()
    PunchingDatas.objects.filter(year=year,month=month,file=file_id).delete()
    messages.success(request,'file deleted succesfully!')
    return redirect(f'/attendance/viewfile/{year}/{month}')

    
@login_required(login_url='loginpage')
@allowed_users(allowed_roles=['admin'])
def punchTableExport(request):
    response = HttpResponse(content_type='text/csv')
    totaldays = [i for i in range(1,32)]
    fname = mnth+"_"+str(yr)+" Punching"
    response['Content-Disposition'] = f'attachment; filename={fname}'+'.csv'
    heads = ['IFID','Name','WFO Attendance','OT Attendance','WFH Attendance',
            'Number of punching mistakes','Night Shift','Deductions','Total Attendance of the month']
    heads.extend(totaldays)
    writer = csv.writer(response)
    writer.writerow(heads)
    
    for x in punchingData:
        writer.writerow([x.ifid ,x.name ,x.WFO_attendance ,x.OT_attendance ,x.WFH_attendance ,x.Punching_Mistake ,
                        x.Night_shift , x.Deduction , x.total_Attendance, 
                        x.date_1 ,x.date_2 ,x.date_3 ,x.date_4 ,x.date_5 ,x.date_6 ,x.date_7 ,
                        x.date_8 ,x.date_9 ,x.date_10 ,x.date_11 ,x.date_12 ,
                        x.date_13 ,x.date_14 ,x.date_15 ,x.date_16 ,x.date_17 ,x.date_18 ,
                        x.date_19 ,x.date_20 ,x.date_21 ,x.date_22 ,x.date_23 ,x.date_24 ,
                        x.date_25 ,x.date_26 ,x.date_27 ,x.date_28 ,x.date_29 ,x.date_30 ,x.date_31 
                        ])
    return response

@login_required(login_url='loginpage')
@allowed_users(allowed_roles=['admin'])
def ifidPunching(request,pk):
    idData = PunchingDatas.objects.get(id=pk)
    ifid = str(idData.ifid).zfill(2)
    idData.ifid = ifid
    mnth = idData.month
    yr = idData.year
    monthid = MonthData.objects.filter(month=mnth, year=yr).first().id 
    
    
    month_num = datetime.datetime.strptime(mnth[:3], '%b').month
    totaldays = monthrange(int(yr), month_num)[1]
    totaldays = [i for i in range(1,totaldays+1)]
    daydata = {}
    new_span = False
    if '-' in mnth:
        month_name = mnth.split("-")[-1]
        new_span = True
        curr_month = mnth.split("-")[-1][:3]
        prev_mnth = mnth[:3]
    else:
        month_name = mnth
        curr_month = mnth[:3]
        prev_mnth = mnth[:3]
    for d in totaldays:
        if d==26:
            daydata[d] = idData.date_26.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(prev_mnth)
        elif d==27:
            daydata[d] = idData.date_27.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(prev_mnth)
        elif d==28:
            daydata[d] = idData.date_28.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(prev_mnth)
        elif d==29:
            daydata[d] = idData.date_29.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(prev_mnth)
        elif d==30:
            daydata[d] = idData.date_30.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(prev_mnth)
        elif d==31:
            daydata[d] = idData.date_31.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(prev_mnth)
        elif d ==1:
            daydata[d] = idData.date_1.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==2:
            daydata[d] = idData.date_2.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==3:
            daydata[d] = idData.date_3.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==4:
            daydata[d] = idData.date_4.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==5:
            daydata[d] = idData.date_5.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==6:
            daydata[d] = idData.date_6.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==7:
            daydata[d] = idData.date_7.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==8:
            daydata[d] = idData.date_8.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==9:
            daydata[d] = idData.date_9.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==10:
            daydata[d] = idData.date_10.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==11:
            daydata[d] = idData.date_11.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==12:
            daydata[d] = idData.date_12.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==13:
            daydata[d] = idData.date_13.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==14:
            daydata[d] = idData.date_14.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==15:
            daydata[d] = idData.date_15.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==16:
            daydata[d] = idData.date_16.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==17:
            daydata[d] = idData.date_17.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==18:
            daydata[d] = idData.date_18.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==19:
            daydata[d] = idData.date_19.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==20:
            daydata[d] = idData.date_20.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            # print(daydata[d][1])
            daydata[d].append(curr_month)
        elif d==21:
            daydata[d] = idData.date_21.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==22:
            daydata[d] = idData.date_22.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==23:
            daydata[d] = idData.date_23.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==24:
            daydata[d] = idData.date_24.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==25:
            daydata[d] = idData.date_25.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        
    # Move 26,27,28..  to 1st::
    if new_span:
        try:
            for i in range(1,26):
                daydata[i] = daydata.pop(i)
        except Exception as er:
            print(er)
    # print(daydata)
    mnletter = mnth[:3]
    # print(daydata)
    context = {'iddata':idData,'month':mnth,'month_name':month_name,'year':yr,'monthid':monthid,'days':daydata,'mnt':mnletter}
    return render(request,'ifid_punching.html',context)



@login_required(login_url='loginpage')
@allowed_users(allowed_roles=['admin'])
def editPunching(request,ifid,year,month):
    edit_punch_data = PunchingDatas.objects.get(ifid=ifid , month=month, year=year)
    # print(edit_punch_data.ifid, edit_punch_data.name)
    form = id_attendance(instance=edit_punch_data)
    if request.method =='POST':
        print(edit_punch_data.ifid, edit_punch_data.name)
        form = id_attendance(request.POST, instance=edit_punch_data)
        if form.is_valid():
            form.save()
            messages.success(request,f' details of IF{ifid} updated succesfully!')
            return redirect(f'/attendance/punchingdetails/{year}/{month}')
        else:
            print('invalid')
            messages.error(request,form.errors)
    return render(request, 'edit_attendance.html',{'form':form, 'month':month,'year':year})
 

# -----------------------------EMPLOYEE LOGIN-------------------------------

@login_required(login_url='loginpage')
def punchingHome(request):
    empname = request.user
    ifid = EmployeeDetails.objects.get(empname=empname).if_id

    if IFMessages.objects.all():
        unreadmsg = EmployeeDetails.objects.get(empname=empname).unread_msg
    else:
        unreadmsg = False
    monthDatas = MonthData.objects.all().order_by('id').reverse()
    mnthdata = []
    dataexist = False
    for mt in monthDatas:
        mnth = mt.month
        if '-' in mnth:
            mt.month_name = mnth.split("-")[-1]
        else:
            mt.month_name = mnth
        yr = mt.year
        id_details = PunchingDatas.objects.filter(ifid =ifid ,year=yr,month=mnth)
        if id_details:
            dataexist = True
            mnthdata.append(mt)
    # if dataexist:
        # print(mnthdata)

    ######## For latest option in template 
    idData = ''
    daydata = {}
    month_name = ''
    if PunchingDatas.objects.filter(ifid=ifid).exists():
        latest = PunchingDatas.objects.filter(ifid=ifid).latest('id')
        print("Latest",latest)
        mon = latest.month
        yr = latest.year
        print("Month : ",mon,yr)
        if "-" in mon:
            month_name = mon.split("-")[-1]
        else:
            month_name = mon
        id_details = PunchingDatas.objects.filter(ifid =ifid ,year=yr,month=month_name)
        idData = PunchingDatas.objects.get(ifid = ifid, month = mon,year=yr)
        month_num = datetime.datetime.strptime(mon[:3], '%b').month
        totaldays = monthrange(int(yr), month_num)[1]
        totaldays = [i for i in range(1,totaldays+1)]
        new_span = False
        if '-' in mon:
            new_span = True
            curr_month = mon.split("-")[-1][:3]
            prev_mnth = mon[:3]
        else:
            curr_month = mon[:3]
            prev_mnth = mon[:3]
        for d in totaldays:
            if d==26:
                daydata[d] = idData.date_26.split("__")
                daydata[d][1]=daydata[d][1].split("\n")
                daydata[d].append(prev_mnth)
            elif d==27:
                daydata[d] = idData.date_27.split("__")
                daydata[d][1]=daydata[d][1].split("\n")
                daydata[d].append(prev_mnth)
            elif d==28:
                daydata[d] = idData.date_28.split("__")
                daydata[d][1]=daydata[d][1].split("\n")
                daydata[d].append(prev_mnth)
            elif d==29:
                daydata[d] = idData.date_29.split("__")
                daydata[d][1]=daydata[d][1].split("\n")
                daydata[d].append(prev_mnth)
            elif d==30:
                daydata[d] = idData.date_30.split("__")
                daydata[d][1]=daydata[d][1].split("\n")
                daydata[d].append(prev_mnth)
            elif d==31:
                daydata[d] = idData.date_31.split("__")
                daydata[d][1]=daydata[d][1].split("\n")
                daydata[d].append(prev_mnth)
            elif d ==1:
                daydata[d] = idData.date_1.split("__")
                daydata[d][1]=daydata[d][1].split("\n")
                daydata[d].append(curr_month)
            elif d==2:
                daydata[d] = idData.date_2.split("__")
                daydata[d][1]=daydata[d][1].split("\n")
                daydata[d].append(curr_month)
            elif d==3:
                daydata[d] = idData.date_3.split("__")
                daydata[d][1]=daydata[d][1].split("\n")
                daydata[d].append(curr_month)
            elif d==4:
                daydata[d] = idData.date_4.split("__")
                daydata[d][1]=daydata[d][1].split("\n")
                daydata[d].append(curr_month)
            elif d==5:
                daydata[d] = idData.date_5.split("__")
                daydata[d][1]=daydata[d][1].split("\n")
                daydata[d].append(curr_month)
            elif d==6:
                daydata[d] = idData.date_6.split("__")
                daydata[d][1]=daydata[d][1].split("\n")
                daydata[d].append(curr_month)
            elif d==7:
                daydata[d] = idData.date_7.split("__")
                daydata[d][1]=daydata[d][1].split("\n")
                daydata[d].append(curr_month)
            elif d==8:
                daydata[d] = idData.date_8.split("__")
                daydata[d][1]=daydata[d][1].split("\n")
                daydata[d].append(curr_month)
            elif d==9:
                daydata[d] = idData.date_9.split("__")
                daydata[d][1]=daydata[d][1].split("\n")
                daydata[d].append(curr_month)
            elif d==10:
                daydata[d] = idData.date_10.split("__")
                daydata[d][1]=daydata[d][1].split("\n")
                daydata[d].append(curr_month)
            elif d==11:
                daydata[d] = idData.date_11.split("__")
                daydata[d][1]=daydata[d][1].split("\n")
                daydata[d].append(curr_month)
            elif d==12:
                daydata[d] = idData.date_12.split("__")
                daydata[d][1]=daydata[d][1].split("\n")
                daydata[d].append(curr_month)
            elif d==13:
                daydata[d] = idData.date_13.split("__")
                daydata[d][1]=daydata[d][1].split("\n")
                daydata[d].append(curr_month)
            elif d==14:
                daydata[d] = idData.date_14.split("__")
                daydata[d][1]=daydata[d][1].split("\n")
                daydata[d].append(curr_month)
            elif d==15:
                daydata[d] = idData.date_15.split("__")
                daydata[d][1]=daydata[d][1].split("\n")
                daydata[d].append(curr_month)
            elif d==16:
                daydata[d] = idData.date_16.split("__")
                daydata[d][1]=daydata[d][1].split("\n")
                daydata[d].append(curr_month)
            elif d==17:
                daydata[d] = idData.date_17.split("__")
                daydata[d][1]=daydata[d][1].split("\n")
                daydata[d].append(curr_month)
            elif d==18:
                daydata[d] = idData.date_18.split("__")
                daydata[d][1]=daydata[d][1].split("\n")
                daydata[d].append(curr_month)
            elif d==19:
                daydata[d] = idData.date_19.split("__")
                daydata[d][1]=daydata[d][1].split("\n")
                daydata[d].append(curr_month)
            elif d==20:
                daydata[d] = idData.date_20.split("__")
                daydata[d][1]=daydata[d][1].split("\n")
                daydata[d].append(curr_month)
            elif d==21:
                daydata[d] = idData.date_21.split("__")
                daydata[d][1]=daydata[d][1].split("\n")
                daydata[d].append(curr_month)
            elif d==22:
                daydata[d] = idData.date_22.split("__")
                daydata[d][1]=daydata[d][1].split("\n")
                daydata[d].append(curr_month)
            elif d==23:
                daydata[d] = idData.date_23.split("__")
                daydata[d][1]=daydata[d][1].split("\n")
                daydata[d].append(curr_month)
            elif d==24:
                daydata[d] = idData.date_24.split("__")
                daydata[d][1]=daydata[d][1].split("\n")
                daydata[d].append(curr_month)
            elif d==25:
                daydata[d] = idData.date_25.split("__")
                daydata[d][1]=daydata[d][1].split("\n")
                daydata[d].append(curr_month)
            
        # Move 26,27,28..  to 1st::
        if new_span:
            try:
                for i in range(1,26):
                    daydata[i] = daydata.pop(i)
            except Exception as er:
                print(er)
    return render(request,'user/att_home.html',{'ifid':ifid,'empname':empname,'punchdata':idData,'days':daydata,'monthdata':mnthdata,'page_title':'INFOLKS | Attendance','unreadmsg':unreadmsg,'nbar':'pnhome','year':yr,'month_name':month_name})


@login_required(login_url='loginpage')
def id_punching(request,year,month):
    empname = request.user
    ifid = EmployeeDetails.objects.get(empname=empname).if_id
    idData = PunchingDatas.objects.get(ifid = ifid, month = month,year=year)
    if "-" in month:
        month_name = month.split("-")[-1]
    else:
        month_name = month
    print(month,year)
    # # For dropdown -------------------------
    monthDatas = MonthData.objects.all().order_by('id').reverse()
    mnthdata = []
    for mt in monthDatas:
        mnth = mt.month
        yr = mt.year
        if '-' in mnth:
            mt.month_name = mnth.split("-")[-1]
        else:
            mt.month_name = mnth
        id_details = PunchingDatas.objects.filter(ifid =ifid ,year=yr,month=mnth)
        if id_details:
            mnthdata.append(mt)

    context = {'punchdata':idData,
            'page_title':'INFOLKS | Attendance', 'month_name':month_name,
            'month':month,'year':year,'ifid':ifid,'empname':empname, 'total_months':mnthdata, 'nbar':'pnhome'}
    return render (request,'user/office_table.html',context)


@login_required(login_url='loginpage')
def id_DayData(request,year,month):
    mnth = month
    yr = year
    empname = request.user
    ifid = EmployeeDetails.objects.get(empname=empname).if_id
    daydata = {}
    mnletter = mnth[:3]
    idData = PunchingDatas.objects.get(ifid = ifid, month = mnth,year=yr)
    month_num = datetime.datetime.strptime(mnth[:3], '%b').month
    totaldays = monthrange(int(yr), month_num)[1]
    totaldays = [i for i in range(1,totaldays+1)]
    # print(mnth,month_num,totaldays)
    new_span = False
    if '-' in mnth:
        new_span = True
        curr_month = mnth.split("-")[-1][:3]
        prev_mnth = mnth[:3]
    else:
        curr_month = mnth[:3]
        prev_mnth = mnth[:3]
    for d in totaldays:
        if d==26:
            daydata[d] = idData.date_26.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(prev_mnth)
        elif d==27:
            daydata[d] = idData.date_27.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(prev_mnth)
        elif d==28:
            daydata[d] = idData.date_28.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(prev_mnth)
        elif d==29:
            daydata[d] = idData.date_29.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(prev_mnth)
        elif d==30:
            daydata[d] = idData.date_30.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(prev_mnth)
        elif d==31:
            daydata[d] = idData.date_31.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(prev_mnth)
        elif d ==1:
            daydata[d] = idData.date_1.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==2:
            daydata[d] = idData.date_2.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==3:
            daydata[d] = idData.date_3.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==4:
            daydata[d] = idData.date_4.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==5:
            daydata[d] = idData.date_5.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==6:
            daydata[d] = idData.date_6.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==7:
            daydata[d] = idData.date_7.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==8:
            daydata[d] = idData.date_8.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==9:
            daydata[d] = idData.date_9.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==10:
            daydata[d] = idData.date_10.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==11:
            daydata[d] = idData.date_11.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==12:
            daydata[d] = idData.date_12.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==13:
            daydata[d] = idData.date_13.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==14:
            daydata[d] = idData.date_14.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==15:
            daydata[d] = idData.date_15.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==16:
            daydata[d] = idData.date_16.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==17:
            daydata[d] = idData.date_17.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==18:
            daydata[d] = idData.date_18.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==19:
            daydata[d] = idData.date_19.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==20:
            daydata[d] = idData.date_20.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==21:
            daydata[d] = idData.date_21.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==22:
            daydata[d] = idData.date_22.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==23:
            daydata[d] = idData.date_23.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==24:
            daydata[d] = idData.date_24.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        elif d==25:
            daydata[d] = idData.date_25.split("__")
            daydata[d][1]=daydata[d][1].split("\n")
            daydata[d].append(curr_month)
        
    # Move 26,27,28..  to 1st::
    if new_span:
        try:
            for i in range(1,26):
                daydata[i] = daydata.pop(i)
        except Exception as er:
            print(er)

    context = {'ifid':ifid,'empname':empname,'month':mnth,'year':yr,'days':daydata,'mnt':mnletter,'page_title' : 'INFOLKS | Attendance','nbar':'pnhome'}
    return render(request,'user/punchings.html',context)